# fbdata.py
import openpyxl
import sys

# Get data from command-line arguments (excluding the script name)
data = sys.argv[1:]

# Try to load an existing Excel workbook or create a new one if it doesn't exist
try:
    wb = openpyxl.load_workbook(filename='fbdata.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Get the active worksheet or create a new one if it doesn't exist
ws = wb.active

# Write headers if the worksheet is empty
if ws.max_row == 1:
    ws['A1'] = 'Insta Page ID'
    ws['B1'] = 'Username'
    ws['C1'] = 'Category'
    ws['D1'] = 'Biography'
    ws['E1'] = 'Caption'
    ws['F1'] = 'Media_url'
    ws['G1'] = 'Timestamp'
# Check if there's enough data for a complete row
if len(data) >= 7:
    # Find the next empty row in the worksheet
    next_row = ws.max_row + 1

    # Write data directly to the Excel sheet
    ws[f'A{next_row}'] = data[0]
    ws[f'B{next_row}'] = data[1]
    ws[f'C{next_row}'] = data[2]
    ws[f'D{next_row}'] = data[3]
    ws[f'E{next_row}'] = data[4]
    ws[f'F{next_row}'] = data[5]
    ws[f'G{next_row}'] = data[6]


    # Save the workbook to a file named 'fbdata.xlsx'
    wb.save('fbdata.xlsx')
else:
    print("Error: Insufficient data for a complete row.")